"""Restaurant admin endpoints — tenant-scoped, auth required."""
import asyncio
import io
from datetime import datetime, timezone, timedelta

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from db import db
from auth import require_restaurant
from whatsapp import notify_order_status
from routes_ws import broadcast as ws_broadcast
from routes_printing import enqueue_order_print
from flemy import emit_flemy_event
from order_security import release_stock
from models import (
    AddonGroupIn, CategoryIn, ProductIn, CouponIn, BannerIn, RestaurantSettings, StatusUpdate,
    ORDER_STATUSES, clean, new_id, now_iso, is_restaurant_open,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])


def rid(user):
    return user["restaurant_id"]


class OrderCycleCloseIn(BaseModel):
    period: str = "day"
    start_date: str | None = None
    end_date: str | None = None
    label: str | None = None


# ---------- restaurant config ----------
@router.get("/restaurant")
async def get_restaurant(user=Depends(require_restaurant)):
    r = await db.restaurants.find_one({"id": rid(user)}, {"_id": 0})
    if not r:
        raise HTTPException(status_code=404, detail="Restaurante não encontrado")
    r["is_open"] = is_restaurant_open(r)
    return r


@router.put("/restaurant")
async def update_restaurant(settings: RestaurantSettings, user=Depends(require_restaurant)):
    updates = {k: v for k, v in settings.model_dump().items() if v is not None}
    if "delivery_fee_mode" in updates and updates.get("delivery_fee_mode") not in ("fixed", "neighborhood"):
        updates["delivery_fee_mode"] = "fixed"
    if "delivery_zones" in updates:
        zones = []
        for zone in updates.get("delivery_zones") or []:
            if not isinstance(zone, dict):
                continue
            name = (zone.get("name") or zone.get("neighborhood") or "").strip()
            fee = float(zone.get("fee") or 0)
            if not name and fee <= 0:
                continue
            zones.append({
                "id": zone.get("id") or new_id(),
                "name": name,
                "neighborhood": (zone.get("neighborhood") or name).strip(),
                "aliases": zone.get("aliases") or "",
                "city_names": zone.get("city_names") or "",
                "cep_prefixes": zone.get("cep_prefixes") or "",
                "fee": max(fee, 0),
                "active": zone.get("active", True) is not False,
            })
        updates["delivery_zones"] = zones
    updates["updated_at"] = now_iso()
    await db.restaurants.update_one({"id": rid(user)}, {"$set": updates})
    r = await db.restaurants.find_one({"id": rid(user)}, {"_id": 0})
    r["is_open"] = is_restaurant_open(r)
    return r


@router.post("/restaurant/toggle-open")
async def toggle_open(user=Depends(require_restaurant)):
    r = await db.restaurants.find_one({"id": rid(user)})
    new_val = not bool(r.get("is_open_manual", True))
    await db.restaurants.update_one({"id": rid(user)}, {"$set": {"is_open_manual": new_val}})
    return {"is_open_manual": new_val}


@router.get("/restaurant/slug")
async def get_restaurant_slug(user=Depends(require_restaurant)):
    r = await db.restaurants.find_one({"id": rid(user)}, {"slug": 1, "_id": 0})
    if not r:
        raise HTTPException(404, "Restaurante não encontrado")
    return {"slug": r.get("slug", "")}


# ---------- categories ----------
@router.get("/categories")
async def list_categories(user=Depends(require_restaurant)):
    return await db.categories.find({"restaurant_id": rid(user)}, {"_id": 0}).sort("sort_order", 1).to_list(500)


@router.post("/categories")
async def create_category(data: CategoryIn, user=Depends(require_restaurant)):
    doc = data.model_dump()
    doc.update({"id": new_id(), "restaurant_id": rid(user), "created_at": now_iso()})
    await db.categories.insert_one(doc)
    return clean(doc)


@router.put("/categories/reorder")
async def reorder_categories(data: dict, user=Depends(require_restaurant)):
    category_ids = data.get("category_ids")
    if not isinstance(category_ids, list) or not category_ids:
        raise HTTPException(status_code=400, detail="Informe a ordem das categorias")

    restaurant_id = rid(user)
    existing = await db.categories.find(
        {"restaurant_id": restaurant_id, "id": {"$in": category_ids}},
        {"id": 1, "_id": 0},
    ).to_list(500)
    existing_ids = {c["id"] for c in existing}
    ordered_ids = [cid for cid in category_ids if cid in existing_ids]

    for index, cid in enumerate(ordered_ids, start=1):
        await db.categories.update_one(
            {"id": cid, "restaurant_id": restaurant_id},
            {"$set": {"sort_order": index, "updated_at": now_iso()}},
        )
    return {"ok": True, "updated": len(ordered_ids)}


@router.put("/categories/{cid}")
async def update_category(cid: str, data: CategoryIn, user=Depends(require_restaurant)):
    res = await db.categories.update_one(
        {"id": cid, "restaurant_id": rid(user)}, {"$set": data.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Categoria não encontrada")
    return await db.categories.find_one({"id": cid, "restaurant_id": rid(user)}, {"_id": 0})


@router.delete("/categories/{cid}")
async def delete_category(cid: str, user=Depends(require_restaurant)):
    await db.categories.delete_one({"id": cid, "restaurant_id": rid(user)})
    await db.products.delete_many({"category_id": cid, "restaurant_id": rid(user)})
    return {"ok": True}


# ---------- products ----------
async def validate_product_suggestions(data: ProductIn, restaurant_id: str, product_id: str = None):
    suggestion_ids = {
        value for value in (data.upsell_product_id, data.downsell_product_id) if value
    }
    if product_id and product_id in suggestion_ids:
        raise HTTPException(status_code=400, detail="O produto nao pode sugerir ele mesmo")
    if not suggestion_ids:
        return
    suggestions = await db.products.find(
        {"restaurant_id": restaurant_id, "id": {"$in": list(suggestion_ids)}},
        {"id": 1, "_id": 0},
    ).to_list(len(suggestion_ids))
    if {product["id"] for product in suggestions} != suggestion_ids:
        raise HTTPException(status_code=400, detail="Produto sugerido invalido")


@router.get("/products")
async def list_products(user=Depends(require_restaurant)):
    return await db.products.find({"restaurant_id": rid(user)}, {"_id": 0}).sort("sort_order", 1).to_list(1000)


@router.post("/products")
async def create_product(data: ProductIn, user=Depends(require_restaurant)):
    await validate_product_suggestions(data, rid(user))
    doc = data.model_dump()
    if doc.get("sort_order", 0) <= 0:
        products = await db.products.find(
            {"restaurant_id": rid(user), "category_id": doc.get("category_id")},
            {"sort_order": 1, "_id": 0},
        ).to_list(1000)
        doc["sort_order"] = max((p.get("sort_order", 0) for p in products), default=0) + 1
    doc.update({"id": new_id(), "restaurant_id": rid(user), "created_at": now_iso()})
    await db.products.insert_one(doc)
    return clean(doc)


@router.put("/products/reorder")
async def reorder_products(data: dict, user=Depends(require_restaurant)):
    product_ids = data.get("product_ids")
    if not isinstance(product_ids, list) or not product_ids:
        raise HTTPException(status_code=400, detail="Informe a ordem dos produtos")

    restaurant_id = rid(user)
    existing = await db.products.find(
        {"restaurant_id": restaurant_id, "id": {"$in": product_ids}},
        {"id": 1, "_id": 0},
    ).to_list(1000)
    existing_ids = {product["id"] for product in existing}
    ordered_ids = [pid for pid in product_ids if pid in existing_ids]

    for index, pid in enumerate(ordered_ids, start=1):
        await db.products.update_one(
            {"id": pid, "restaurant_id": restaurant_id},
            {"$set": {"sort_order": index, "updated_at": now_iso()}},
        )
    return {"ok": True, "updated": len(ordered_ids)}


@router.put("/products/{pid}")
async def update_product(pid: str, data: ProductIn, user=Depends(require_restaurant)):
    await validate_product_suggestions(data, rid(user), pid)
    res = await db.products.update_one(
        {"id": pid, "restaurant_id": rid(user)}, {"$set": data.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    return await db.products.find_one({"id": pid, "restaurant_id": rid(user)}, {"_id": 0})


@router.delete("/products/{pid}")
async def delete_product(pid: str, user=Depends(require_restaurant)):
    await db.products.delete_one({"id": pid, "restaurant_id": rid(user)})
    await db.addon_groups.update_many(
        {"restaurant_id": rid(user), "product_ids": {"$in": [pid]}},
        {"$pull": {"product_ids": pid}},
    )
    return {"ok": True}


# ---------- reusable addon groups ----------
async def validate_addon_products(product_ids: list[str], restaurant_id: str):
    clean_ids = [pid for pid in dict.fromkeys(product_ids or []) if pid]
    if not clean_ids:
        return []
    products = await db.products.find(
        {"restaurant_id": restaurant_id, "id": {"$in": clean_ids}},
        {"id": 1, "_id": 0},
    ).to_list(len(clean_ids))
    found = {p["id"] for p in products}
    missing = [pid for pid in clean_ids if pid not in found]
    if missing:
        raise HTTPException(status_code=400, detail="Produto invalido na lista de adicionais")
    return clean_ids


@router.get("/addon-groups")
async def list_addon_groups(user=Depends(require_restaurant)):
    return await db.addon_groups.find(
        {"restaurant_id": rid(user)}, {"_id": 0}
    ).sort("sort_order", 1).to_list(500)


@router.post("/addon-groups")
async def create_addon_group(data: AddonGroupIn, user=Depends(require_restaurant)):
    restaurant_id = rid(user)
    doc = data.model_dump()
    doc["product_ids"] = await validate_addon_products(doc.get("product_ids") or [], restaurant_id)
    if doc.get("sort_order", 0) <= 0:
        groups = await db.addon_groups.find(
            {"restaurant_id": restaurant_id}, {"sort_order": 1, "_id": 0}
        ).to_list(500)
        doc["sort_order"] = max((g.get("sort_order", 0) for g in groups), default=0) + 1
    doc.update({"id": new_id(), "restaurant_id": restaurant_id, "created_at": now_iso()})
    await db.addon_groups.insert_one(doc)
    return clean(doc)


@router.put("/addon-groups/{gid}")
async def update_addon_group(gid: str, data: AddonGroupIn, user=Depends(require_restaurant)):
    restaurant_id = rid(user)
    patch = data.model_dump()
    patch["id"] = gid
    patch["product_ids"] = await validate_addon_products(patch.get("product_ids") or [], restaurant_id)
    patch["updated_at"] = now_iso()
    res = await db.addon_groups.update_one(
        {"id": gid, "restaurant_id": restaurant_id},
        {"$set": patch},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Grupo de adicionais nao encontrado")
    return await db.addon_groups.find_one({"id": gid, "restaurant_id": restaurant_id}, {"_id": 0})


@router.delete("/addon-groups/{gid}")
async def delete_addon_group(gid: str, user=Depends(require_restaurant)):
    await db.addon_groups.delete_one({"id": gid, "restaurant_id": rid(user)})
    return {"ok": True}


@router.get("/products/export")
async def export_products(user=Depends(require_restaurant)):
    products = await db.products.find({"restaurant_id": rid(user)}, {"_id": 0}).sort("sort_order", 1).to_list(1000)

    # Build category id -> name map
    cat_ids = list({p.get("category_id") for p in products if p.get("category_id")})
    categories = await db.categories.find({"id": {"$in": cat_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
    cat_map = {c["id"]: c["name"] for c in categories}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"
    headers = [
        "name", "description", "price", "promotional_price",
        "category_name", "is_available", "is_best_seller", "is_featured",
        "track_stock", "stock_quantity", "image_url",
    ]
    ws.append(headers)

    for p in products:
        ws.append([
            p.get("name", ""),
            p.get("description", ""),
            p.get("price", 0),
            p.get("promotional_price", ""),
            cat_map.get(p.get("category_id", ""), ""),
            p.get("is_available", True),
            p.get("is_best_seller", False),
            p.get("is_featured", False),
            p.get("track_stock", False),
            p.get("stock_quantity", ""),
            p.get("image_url", ""),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=produtos.xlsx"},
    )


@router.post("/products/import")
async def import_products(file: UploadFile = File(...), user=Depends(require_restaurant)):
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Arquivo inválido: {exc}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "updated": 0, "errors": []}

    # Determine column positions from header row
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}

    required_cols = {"name", "price"}
    missing = required_cols - set(col.keys())
    if missing:
        raise HTTPException(400, f"Colunas obrigatórias ausentes: {missing}")

    restaurant_id = rid(user)
    imported = 0
    updated = 0
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            def _get(field, default=None):
                idx = col.get(field)
                if idx is None:
                    return default
                val = row[idx] if idx < len(row) else None
                return val if val is not None else default

            name = str(_get("name", "")).strip()
            if not name:
                errors.append(f"Linha {row_num}: nome vazio, ignorado")
                continue

            try:
                price = float(_get("price", 0) or 0)
            except (ValueError, TypeError):
                errors.append(f"Linha {row_num}: preço inválido para '{name}'")
                continue

            promo_raw = _get("promotional_price")
            try:
                promotional_price = float(promo_raw) if promo_raw not in (None, "", "none", "null") else None
            except (ValueError, TypeError):
                promotional_price = None

            description = str(_get("description", "") or "")
            category_name = str(_get("category_name", "") or "").strip()
            is_available = bool(_get("is_available", True))
            is_best_seller = bool(_get("is_best_seller", False))

            # Resolve or create category
            category_id = None
            if category_name:
                cat = await db.categories.find_one(
                    {"restaurant_id": restaurant_id, "name": {"$regex": f"^{category_name}$", "$options": "i"}}
                )
                if cat:
                    category_id = cat["id"]
                else:
                    category_id = new_id()
                    await db.categories.insert_one({
                        "id": category_id,
                        "restaurant_id": restaurant_id,
                        "name": category_name,
                        "sort_order": 0,
                        "created_at": now_iso(),
                    })

            payload = {
                "name": name,
                "description": description,
                "price": price,
                "promotional_price": promotional_price,
                "category_id": category_id,
                "is_available": is_available,
                "is_best_seller": is_best_seller,
                "updated_at": now_iso(),
            }

            existing = await db.products.find_one(
                {"restaurant_id": restaurant_id, "name": {"$regex": f"^{name}$", "$options": "i"}}
            )
            if existing:
                await db.products.update_one(
                    {"id": existing["id"], "restaurant_id": restaurant_id},
                    {"$set": payload},
                )
                updated += 1
            else:
                payload.update({
                    "id": new_id(),
                    "restaurant_id": restaurant_id,
                    "is_featured": False,
                    "track_stock": False,
                    "stock_quantity": None,
                    "image_url": None,
                    "sort_order": 0,
                    "created_at": now_iso(),
                })
                await db.products.insert_one(payload)
                imported += 1

        except Exception as exc:
            errors.append(f"Linha {row_num}: erro inesperado — {exc}")

    return {"imported": imported, "updated": updated, "errors": errors}


# ---------- orders ----------
TERMINAL_ORDER_STATUSES = {"completed", "cancelled"}


def _local_date_to_utc_iso(value: str | None, end_of_day: bool = False):
    if not value:
        return None
    try:
        from zoneinfo import ZoneInfo
        local_tz = ZoneInfo("America/Sao_Paulo")
    except Exception:
        local_tz = timezone.utc

    raw = value.strip()
    try:
        if len(raw) == 10:
            dt = datetime.strptime(raw, "%Y-%m-%d")
            if end_of_day:
                dt = dt.replace(hour=23, minute=59, second=59, microsecond=999000)
            else:
                dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
            dt = dt.replace(tzinfo=local_tz)
        else:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=local_tz)
        return dt.astimezone(timezone.utc).isoformat()
    except Exception:
        raise HTTPException(status_code=400, detail="Data invalida")


def _order_query(
    restaurant_id: str,
    status: str | None = None,
    payment_status: str | None = None,
    source: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    cycle: str = "current",
):
    conditions = [{"restaurant_id": restaurant_id}]

    if status and status != "all":
        conditions.append({"status": status})
    if payment_status and payment_status != "all":
        conditions.append({"payment_status": payment_status})
    if source and source != "all":
        if source == "online":
            conditions.append({"$or": [{"source": {"$exists": False}}, {"source": {"$in": ["online", "web", "public"]}}]})
        else:
            conditions.append({"source": source})
    if cycle == "current":
        conditions.append({"$or": [{"cycle_id": {"$exists": False}}, {"cycle_id": None}]})
    elif cycle == "history":
        conditions.append({"cycle_id": {"$exists": True, "$ne": None}})

    created = {}
    start_iso = _local_date_to_utc_iso(start_date, False)
    end_iso = _local_date_to_utc_iso(end_date, True)
    if start_iso:
        created["$gte"] = start_iso
    if end_iso:
        created["$lte"] = end_iso
    if created:
        conditions.append({"created_at": created})

    return {"$and": conditions} if len(conditions) > 1 else conditions[0]


def _search_order(order, text: str | None):
    if not text:
        return True
    q = text.strip().lower()
    if not q:
        return True
    customer = order.get("customer") or {}
    haystack = " ".join([
        str(order.get("order_number") or ""),
        str(order.get("id") or ""),
        str(customer.get("name") or ""),
        str(customer.get("phone") or ""),
        str(order.get("payment_method") or ""),
    ]).lower()
    return q in haystack


async def _fetch_orders_for_admin(
    restaurant_id: str,
    status: str | None = None,
    payment_status: str | None = None,
    source: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    cycle: str = "current",
    search: str | None = None,
    limit: int = 500,
):
    max_limit = min(max(limit or 500, 1), 10000)
    query = _order_query(restaurant_id, status, payment_status, source, start_date, end_date, cycle)
    fetch_limit = 10000 if search else max_limit
    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(fetch_limit)
    if search:
        orders = [o for o in orders if _search_order(o, search)][:max_limit]
    return orders


def _source_label(order):
    source = order.get("source")
    if source == "pdv":
        return "PDV"
    if source == "whatsapp":
        return "WhatsApp"
    if source == "table_qr":
        return "Mesa / QR"
    return "Online"


def _build_orders_summary(orders):
    valid_orders = [o for o in orders if o.get("status") != "cancelled"]
    revenue = round(sum(float(o.get("total") or 0) for o in valid_orders), 2)
    total_orders = len(orders)
    by_status = {key: 0 for key in ORDER_STATUSES}
    by_payment = {}
    by_source = {}
    by_day = {}

    for order in orders:
        status = order.get("status") or "pending"
        by_status[status] = by_status.get(status, 0) + 1
        if status != "cancelled":
            total = float(order.get("total") or 0)
            payment = order.get("payment_method") or "outros"
            by_payment[payment] = round(by_payment.get(payment, 0) + total, 2)
            source = _source_label(order)
            by_source[source] = round(by_source.get(source, 0) + total, 2)
            day = (order.get("created_at") or "")[:10] or "sem-data"
            if day not in by_day:
                by_day[day] = {"date": day, "orders": 0, "revenue": 0.0}
            by_day[day]["orders"] += 1
            by_day[day]["revenue"] = round(by_day[day]["revenue"] + total, 2)

    return {
        "summary": {
            "orders": total_orders,
            "valid_orders": len(valid_orders),
            "revenue": revenue,
            "avg_ticket": round(revenue / len(valid_orders), 2) if valid_orders else 0,
            "pending": by_status.get("pending", 0),
            "in_progress": sum(by_status.get(s, 0) for s in ["accepted", "preparing", "ready", "out_for_delivery"]),
            "cancelled": by_status.get("cancelled", 0),
        },
        "by_status": by_status,
        "by_payment": by_payment,
        "by_source": by_source,
        "by_day": sorted(by_day.values(), key=lambda item: item["date"]),
    }


@router.get("/orders")
async def list_orders(
    status: str = None,
    payment_status: str = None,
    source: str = None,
    start_date: str = None,
    end_date: str = None,
    cycle: str = "current",
    search: str = None,
    limit: int = 500,
    user=Depends(require_restaurant),
):
    return await _fetch_orders_for_admin(
        rid(user), status, payment_status, source, start_date, end_date, cycle, search, limit
    )


@router.get("/orders/summary")
async def orders_summary(
    status: str = None,
    payment_status: str = None,
    source: str = None,
    start_date: str = None,
    end_date: str = None,
    cycle: str = "current",
    search: str = None,
    user=Depends(require_restaurant),
):
    orders = await _fetch_orders_for_admin(
        rid(user), status, payment_status, source, start_date, end_date, cycle, search, 10000
    )
    last_cycle = await db.order_cycles.find_one(
        {"restaurant_id": rid(user)}, {"_id": 0}, sort=[("closed_at", -1)]
    )
    data = _build_orders_summary(orders)
    data["last_cycle"] = last_cycle
    data["filters"] = {
        "status": status or "all",
        "payment_status": payment_status or "all",
        "source": source or "all",
        "start_date": start_date,
        "end_date": end_date,
        "cycle": cycle,
    }
    return data


@router.get("/orders/export")
async def export_orders(
    status: str = None,
    payment_status: str = None,
    source: str = None,
    start_date: str = None,
    end_date: str = None,
    cycle: str = "current",
    search: str = None,
    user=Depends(require_restaurant),
):
    orders = await _fetch_orders_for_admin(
        rid(user), status, payment_status, source, start_date, end_date, cycle, search, 10000
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    ws.append([
        "Numero", "Criado em", "Cliente", "Telefone", "Status", "Canal",
        "Tipo", "Pagamento", "Subtotal", "Entrega", "Desconto", "Total",
        "Itens", "Ciclo",
    ])
    for order in orders:
        customer = order.get("customer") or {}
        items = "; ".join(
            f"{item.get('quantity', 1)}x {item.get('product_name', '')}"
            for item in order.get("items", [])
        )
        ws.append([
            order.get("order_number"),
            order.get("created_at"),
            customer.get("name"),
            customer.get("phone"),
            order.get("status"),
            _source_label(order),
            order.get("type"),
            order.get("payment_method"),
            float(order.get("subtotal") or 0),
            float(order.get("delivery_fee") or 0),
            float(order.get("discount") or 0),
            float(order.get("total") or 0),
            items,
            order.get("cycle_label") or order.get("cycle_id") or "",
        ])
    for col in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 44)
        ws.column_dimensions[col[0].column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pedidos.xlsx"},
    )


@router.get("/order-cycles")
async def list_order_cycles(user=Depends(require_restaurant)):
    return await db.order_cycles.find(
        {"restaurant_id": rid(user)}, {"_id": 0}
    ).sort("closed_at", -1).to_list(200)


def _cycle_bounds(period: str, start_date: str | None, end_date: str | None):
    try:
        from zoneinfo import ZoneInfo
        local_tz = ZoneInfo("America/Sao_Paulo")
    except Exception:
        local_tz = timezone.utc

    now_local = datetime.now(local_tz)
    if start_date or end_date:
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="Informe data inicial e final")
        return _local_date_to_utc_iso(start_date, False), _local_date_to_utc_iso(end_date, True)
    if period == "week":
        start_local = now_local.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=now_local.weekday())
    else:
        start_local = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
    return start_local.astimezone(timezone.utc).isoformat(), datetime.now(timezone.utc).isoformat()


def _current_cycle_condition():
    return {"$or": [{"cycle_id": {"$exists": False}}, {"cycle_id": None}]}


def _date_range_condition(start_iso: str, end_iso: str):
    return {"created_at": {"$gte": start_iso, "$lte": end_iso}}


def _cycle_orders_query(restaurant_id: str, start_iso: str | None = None, end_iso: str | None = None):
    conditions = [
        {"restaurant_id": restaurant_id},
        _current_cycle_condition(),
    ]
    if start_iso and end_iso:
        conditions.append(_date_range_condition(start_iso, end_iso))
    return {"$and": conditions}


def _with_terminal_status(query: dict):
    return {"$and": [query, {"status": {"$in": list(TERMINAL_ORDER_STATUSES)}}]}


def _with_open_status(query: dict):
    return {"$and": [query, {"status": {"$nin": list(TERMINAL_ORDER_STATUSES)}}]}


@router.post("/order-cycles/close")
async def close_order_cycle(data: OrderCycleCloseIn, user=Depends(require_restaurant)):
    period = data.period if data.period in {"day", "week", "custom"} else "day"
    start_iso, end_iso = _cycle_bounds(period, data.start_date, data.end_date)
    restaurant_id = rid(user)
    base_query = _cycle_orders_query(restaurant_id, start_iso, end_iso)
    closable_query = _with_terminal_status(base_query)
    orders = await db.orders.find(closable_query, {"_id": 0}).to_list(10000)
    open_orders = await db.orders.count_documents(_with_open_status(base_query))

    # Alguns pedidos antigos foram gravados com created_at fora do recorte local/UTC
    # esperado. Se a tela mostra finalizados no ciclo atual, o fechamento deve limpar
    # esses pedidos em vez de travar com falso "nenhum pedido encontrado".
    if not orders and period != "custom" and not data.start_date and not data.end_date:
        base_query = _cycle_orders_query(restaurant_id)
        closable_query = _with_terminal_status(base_query)
        orders = await db.orders.find(closable_query, {"_id": 0}).to_list(10000)
        open_orders = await db.orders.count_documents(_with_open_status(base_query))

    if not orders:
        raise HTTPException(
            status_code=400,
            detail="Nenhum pedido finalizado ou cancelado encontrado para fechar neste ciclo",
        )

    summary = _build_orders_summary(orders)["summary"]
    closed_at = now_iso()
    label = data.label or ("Fechamento diario" if period == "day" else "Fechamento semanal" if period == "week" else "Fechamento personalizado")
    doc = {
        "id": new_id(),
        "restaurant_id": restaurant_id,
        "period": period,
        "label": label,
        "started_at": start_iso,
        "ended_at": end_iso,
        "closed_at": closed_at,
        "orders_count": summary["orders"],
        "valid_orders": summary["valid_orders"],
        "cancelled_orders": summary["cancelled"],
        "revenue": summary["revenue"],
        "avg_ticket": summary["avg_ticket"],
        "open_orders_left": open_orders,
        "created_by": user.get("id"),
    }
    await db.order_cycles.insert_one(doc)
    await db.orders.update_many(
        {"id": {"$in": [o["id"] for o in orders]}, "restaurant_id": restaurant_id},
        {"$set": {"cycle_id": doc["id"], "cycle_label": label, "cycle_closed_at": closed_at, "updated_at": closed_at}},
    )
    return clean(doc)


@router.post("/order-cycles/{cycle_id}/reopen")
async def reopen_order_cycle(cycle_id: str, user=Depends(require_restaurant)):
    cycle = await db.order_cycles.find_one({"id": cycle_id, "restaurant_id": rid(user)}, {"_id": 0})
    if not cycle:
        raise HTTPException(status_code=404, detail="Ciclo nao encontrado")
    await db.orders.update_many(
        {"restaurant_id": rid(user), "cycle_id": cycle_id},
        {"$unset": {"cycle_id": "", "cycle_label": "", "cycle_closed_at": ""}, "$set": {"updated_at": now_iso()}},
    )
    await db.order_cycles.delete_one({"id": cycle_id, "restaurant_id": rid(user)})
    return {"ok": True}


@router.get("/orders/{oid}")
async def get_order(oid: str, user=Depends(require_restaurant)):
    o = await db.orders.find_one({"id": oid, "restaurant_id": rid(user)}, {"_id": 0})
    if not o:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    return o


@router.put("/orders/{oid}/status")
async def update_order_status(oid: str, data: StatusUpdate, user=Depends(require_restaurant)):
    if data.status not in ORDER_STATUSES:
        raise HTTPException(status_code=400, detail="Status inválido")
    previous_order = await db.orders.find_one({"id": oid, "restaurant_id": rid(user)}, {"_id": 0})
    if not previous_order:
        raise HTTPException(status_code=404, detail="Pedido nao encontrado")
    updates = {"status": data.status, "updated_at": now_iso()}
    if data.status == "accepted" and previous_order.get("status") != "accepted":
        updates["accepted_at"] = now_iso()
    if data.status == "out_for_delivery" and data.delivery_person_id:
        delivery_person = await db.delivery_people.find_one(
            {"id": data.delivery_person_id, "restaurant_id": rid(user), "is_active": True},
            {"_id": 0},
        )
        if not delivery_person:
            raise HTTPException(status_code=400, detail="Entregador invalido")
        updates["delivery_person"] = {
            "id": delivery_person["id"],
            "name": delivery_person.get("name"),
            "phone": delivery_person.get("phone"),
            "vehicle_type": delivery_person.get("vehicle_type"),
            "vehicle_plate": delivery_person.get("vehicle_plate"),
            "delivery_fee": delivery_person.get("delivery_fee", 0),
        }
        updates["delivery_assigned_at"] = now_iso()
    res = await db.orders.update_one(
        {"id": oid, "restaurant_id": rid(user)},
        {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pedido não encontrado")
    order = await db.orders.find_one({"id": oid, "restaurant_id": rid(user)}, {"_id": 0})
    if (
        data.status == "cancelled"
        and previous_order.get("status") in {"pending", "accepted"}
        and order.get("stock_reservations")
        and not order.get("stock_released_at")
    ):
        await release_stock(db, order["restaurant_id"], order["stock_reservations"])
        await db.orders.update_one(
            {"id": oid, "restaurant_id": rid(user)},
            {"$set": {"stock_released_at": now_iso()}},
        )
        order["stock_released_at"] = now_iso()
    if order.get("table_id") and data.status in {"completed", "cancelled"}:
        open_table_orders = await db.orders.count_documents({
            "restaurant_id": order["restaurant_id"],
            "table_id": order["table_id"],
            "status": {"$nin": ["completed", "cancelled"]},
        })
        if open_table_orders == 0:
            await db.tables.update_one(
                {"id": order["table_id"], "restaurant_id": order["restaurant_id"]},
                {"$set": {"status": "available", "updated_at": now_iso()}},
            )
    # Fire-and-forget WhatsApp notification
    asyncio.create_task(notify_order_status(order, data.status))
    asyncio.create_task(ws_broadcast(order['restaurant_id'], 'order_updated', {'id': order['id'], 'status': data.status}))
    restaurant = await db.restaurants.find_one({"id": order["restaurant_id"]}, {"_id": 0})
    if data.status == "accepted":
        await enqueue_order_print(restaurant, order, "accepted")
    event = "order.cancelled" if data.status == "cancelled" else "order.status_changed"
    asyncio.create_task(emit_flemy_event(restaurant, event, order, {"new_status": data.status}))
    return order


# ---------- coupons ----------
@router.get("/coupons")
async def list_coupons(user=Depends(require_restaurant)):
    return await db.coupons.find({"restaurant_id": rid(user)}, {"_id": 0}).to_list(200)


@router.post("/coupons")
async def create_coupon(data: CouponIn, user=Depends(require_restaurant)):
    doc = data.model_dump()
    doc["code"] = doc["code"].upper()
    doc.update({"id": new_id(), "restaurant_id": rid(user), "used_count": 0, "created_at": now_iso()})
    await db.coupons.insert_one(doc)
    return clean(doc)


@router.put("/coupons/{cid}")
async def update_coupon(cid: str, data: CouponIn, user=Depends(require_restaurant)):
    payload = data.model_dump()
    payload["code"] = payload["code"].upper()
    res = await db.coupons.update_one({"id": cid, "restaurant_id": rid(user)}, {"$set": payload})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cupom não encontrado")
    return await db.coupons.find_one({"id": cid, "restaurant_id": rid(user)}, {"_id": 0})


@router.delete("/coupons/{cid}")
async def delete_coupon(cid: str, user=Depends(require_restaurant)):
    await db.coupons.delete_one({"id": cid, "restaurant_id": rid(user)})
    return {"ok": True}


# ---------- banners ----------
@router.get("/banners")
async def list_banners(user=Depends(require_restaurant)):
    return await db.banners.find({"restaurant_id": rid(user)}, {"_id": 0}).sort("sort_order", 1).to_list(100)


@router.post("/banners")
async def create_banner(data: BannerIn, user=Depends(require_restaurant)):
    doc = data.model_dump()
    if doc.get("sort_order", 0) <= 0:
        banners = await db.banners.find(
            {"restaurant_id": rid(user)}, {"sort_order": 1, "_id": 0}
        ).to_list(100)
        doc["sort_order"] = max((b.get("sort_order", 0) for b in banners), default=0) + 1
    doc.update({"id": new_id(), "restaurant_id": rid(user), "created_at": now_iso()})
    await db.banners.insert_one(doc)
    return clean(doc)


@router.put("/banners/reorder")
async def reorder_banners(data: dict, user=Depends(require_restaurant)):
    banner_ids = data.get("banner_ids")
    if not isinstance(banner_ids, list) or not banner_ids:
        raise HTTPException(status_code=400, detail="Informe a ordem dos banners")

    restaurant_id = rid(user)
    existing = await db.banners.find(
        {"restaurant_id": restaurant_id, "id": {"$in": banner_ids}},
        {"id": 1, "_id": 0},
    ).to_list(100)
    existing_ids = {banner["id"] for banner in existing}
    ordered_ids = [bid for bid in banner_ids if bid in existing_ids]

    for index, bid in enumerate(ordered_ids, start=1):
        await db.banners.update_one(
            {"id": bid, "restaurant_id": restaurant_id},
            {"$set": {"sort_order": index, "updated_at": now_iso()}},
        )
    return {"ok": True, "updated": len(ordered_ids)}


@router.put("/banners/{bid}")
async def update_banner(bid: str, data: BannerIn, user=Depends(require_restaurant)):
    res = await db.banners.update_one({"id": bid, "restaurant_id": rid(user)}, {"$set": data.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Banner não encontrado")
    return await db.banners.find_one({"id": bid, "restaurant_id": rid(user)}, {"_id": 0})


@router.delete("/banners/{bid}")
async def delete_banner(bid: str, user=Depends(require_restaurant)):
    await db.banners.delete_one({"id": bid, "restaurant_id": rid(user)})
    return {"ok": True}


# ---------- dashboard & reports ----------
def _today_start_iso():
    try:
        from zoneinfo import ZoneInfo
        now = datetime.now(ZoneInfo("America/Sao_Paulo"))
    except Exception:
        now = datetime.now()
    return now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()


@router.get("/dashboard")
async def dashboard(user=Depends(require_restaurant)):
    r = await db.restaurants.find_one({"id": rid(user)})
    orders = await db.orders.find({"restaurant_id": rid(user)}, {"_id": 0}).to_list(5000)
    today = _today_start_iso()[:10]
    today_orders = [o for o in orders if (o.get("created_at") or "")[:10] == today and o.get("status") != "cancelled"]
    revenue = sum(o["total"] for o in today_orders)
    in_progress = [o for o in orders if o.get("status") in ("pending", "accepted", "preparing", "ready", "out_for_delivery")]

    # top products
    counter = {}
    for o in orders:
        if o.get("status") == "cancelled":
            continue
        for it in o.get("items", []):
            counter[it["product_name"]] = counter.get(it["product_name"], 0) + it.get("quantity", 1)
    top = sorted(counter.items(), key=lambda x: x[1], reverse=True)[:5]

    return {
        "orders_today": len(today_orders),
        "revenue_today": round(revenue, 2),
        "avg_ticket": round(revenue / len(today_orders), 2) if today_orders else 0,
        "in_progress": len(in_progress),
        "is_open": is_restaurant_open(r) if r else False,
        "is_open_manual": r.get("is_open_manual", True) if r else True,
        "top_products": [{"name": n, "qty": q} for n, q in top],
        "recent_orders": sorted(orders, key=lambda o: o.get("created_at", ""), reverse=True)[:8],
    }


@router.get("/reports")
async def reports(period: str = "7d", user=Depends(require_restaurant)):
    days = {"today": 1, "7d": 7, "30d": 30, "90d": 90}.get(period, 7)
    try:
        from zoneinfo import ZoneInfo
        now = datetime.now(ZoneInfo("America/Sao_Paulo"))
    except Exception:
        now = datetime.now()
    start = (now - timedelta(days=days)).isoformat()
    # Fetch ALL orders (including cancelled) so frontend can compute cancel rate
    all_orders = await db.orders.find(
        {"restaurant_id": rid(user), "created_at": {"$gte": start}},
        {"_id": 0},
    ).to_list(5000)
    return {"orders": all_orders}
